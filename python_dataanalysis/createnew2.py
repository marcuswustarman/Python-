import openpyxl
import random
from copy import copy

def copy_cell(cell, new_cell):
    """完整复制单元格内容和格式"""
    # 复制值（读取公式计算结果）
    new_cell.value = cell.value
    
    # 复制样式
    new_cell.font = copy(cell.font)
    new_cell.border = copy(cell.border)
    new_cell.fill = copy(cell.fill)
    new_cell.number_format = copy(cell.number_format)
    new_cell.alignment = copy(cell.alignment)

# 文件设置
input_file = "根据留存率推算用户增长的公式.xlsx"
output_file = "new2根据留存率推算用户增长的公式.xlsx"

# 读取原文件（关键设置：data_only=True）
wb = openpyxl.load_workbook(input_file, data_only=True)
sheet = wb["日期表"]

# 创建新工作簿
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
new_sheet.title = "日期表"

# 全表复制（仅值）
for row in sheet.iter_rows():
    for cell in row:
        new_cell = new_sheet[cell.coordinate]
        copy_cell(cell, new_cell)

# 处理指定区域
for row in range(4, 36):    # Excel行号4-35
    for col in range(5, 37): # E(5)到AJ(36)
        cell = new_sheet.cell(row=row, column=col)
        
        # 仅处理数值内容
        if isinstance(cell.value, (int, float)):
            # 生成随机系数并计算
            ratio = random.uniform(0.8, 1.2)
            cell.value = round(cell.value * ratio)

# 保存文件
new_wb.save(output_file)