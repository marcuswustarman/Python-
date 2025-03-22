import openpyxl
from openpyxl.utils import get_column_letter
import shutil
import os

def shift_row(row_data):
    """将行数据中的非零值左移，零值右置，保留原顺序"""
    non_zeros = [cell.value for cell in row_data if cell.value != 0]
    zeros_count = len(row_data) - len(non_zeros)
    return non_zeros + [0] * zeros_count

def process_excel_with_format():
    try:
        # 文件路径配置
        src_file = "new2根据留存率推算用户增长的公式.xlsx"
        dest_file = "new3根据留存率推算用户增长的公式.xlsx"
        
        if not os.path.exists(src_file):
            raise FileNotFoundError(f"源文件 {src_file} 不存在")

        # 复制原文件（保留所有格式）
        shutil.copyfile(src_file, dest_file)

        # 加载目标工作簿
        wb = openpyxl.load_workbook(dest_file)
        if "日期表" not in wb.sheetnames:
            raise KeyError("原文件中不存在「日期表」工作表")

        # 处理原日期表
        src_sheet = wb["日期表"]
        
        # 定义处理范围 F5-AJ35（openpyxl行列从1开始）
        start_row, end_row = 5, 35
        start_col, end_col = 6, 36  # F=6, AJ=36

        # 逐行处理
        for row_idx in range(start_row, end_row + 1):
            row_data = [src_sheet.cell(row=row_idx, column=col) for col in range(start_col, end_col + 1)]
            processed_values = shift_row(row_data)
            
            # 重新写入并保留格式
            for col_offset, value in enumerate(processed_values):
                cell = src_sheet.cell(row=row_idx, column=start_col + col_offset)
                cell.value = value  # 只修改值，保留原有样式

        # 创建处理后的新sheet（复制原表结构和格式）
        new_sheet = wb.copy_worksheet(src_sheet)
        new_sheet.title = "处理后的日期表"

        # 保存修改
        wb.save(dest_file)
        print("处理成功！输出文件:", dest_file)

    except Exception as e:
        print(f"错误发生: {str(e)}")
        if os.path.exists(dest_file):
            os.remove(dest_file)
            print("已清理临时文件")

if __name__ == "__main__":
    process_excel_with_format()